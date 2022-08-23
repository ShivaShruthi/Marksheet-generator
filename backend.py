import os
from numpy import NAN, nan
import pandas as pd 
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from flask_mail import Mail,Message
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


#border function
def set_border(ws, cell_range):
    thin = Side(border_style="thin")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)



def generatemarksheet(p,n):

    current_path = os.getcwd()
    response_path= current_path+'/sample_input'
    os.chdir(response_path)


    Responses_df = pd.read_csv('responses.csv',index_col='Timestamp')

    roll_list = Responses_df["Roll Number"].values.tolist()
    names_list=Responses_df["Name"].values.tolist()

    

    st_roll_list_df=pd.read_csv('master_roll.csv')
    st_roll_list = st_roll_list_df["roll"].values.tolist()
    names_list=st_roll_list_df["name"].values.tolist()
    

    dict={}
    for roll in st_roll_list:
        for name in names_list:
            dict[roll]=name

    crt_options = []
    crt_options = Responses_df.loc[Responses_df['Roll Number']=='ANSWER'].values.tolist()
    if(len(crt_options)==0):
        return True
    crt_options=crt_options[0] 
    crt_options=crt_options[6:]

    os.chdir(current_path)
    if not os.path.exists("my_output"):
        os.mkdir('my_output')

    st_no = 0
    st_no_master=0
    while(st_no_master < len(st_roll_list) ):    

        wb = openpyxl.Workbook()
        sheet=wb.active
        img = openpyxl.drawing.image.Image('pic.png')
        img.width =613.3
        img.height = 82.5
        img.anchor='A1'
        sheet.add_image(img,'A1')
        sheet.column_dimensions['A'].width = 17
        sheet.column_dimensions['B'].width = 17
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['E'].width = 17

        sheet.merge_cells('A5:E5')
        sheet.cell(row = 5, column = 1).font = Font(size = 18, bold = True,name='Century',underline='single')
        sheet.cell(row = 5, column = 1).alignment= Alignment('center')
        sheet.cell(row = 5, column = 1).value='Mark Sheet'

        sheet.cell(row = 6, column = 1).alignment= Alignment('right')
        sheet.cell(row = 6, column = 1).font = Font(size = 12,name='Century')
        sheet.cell(row = 6, column = 1).value='Name:'

        sheet.merge_cells('B6:C6')
        sheet.cell(row = 6, column = 2).alignment= Alignment('left')
        sheet.cell(row = 6, column = 2).font = Font(size = 12,bold=True,name='Century')
        sheet.cell(row = 6, column = 2).value= dict[st_roll_list[st_no_master]]

        sheet.cell(row = 6, column = 4).alignment= Alignment('right')
        sheet.cell(row = 6, column = 4).font = Font(size = 12,name='Century')
        sheet.cell(row = 6, column = 4).value='Exam'

        sheet.cell(row = 6, column = 5).alignment= Alignment('left')
        sheet.cell(row = 6, column = 5).font = Font(size = 12,bold=True,name='Century')
        sheet.cell(row = 6, column = 5).value='quiz'

        sheet.cell(row = 7, column = 1).alignment= Alignment('right')
        sheet.cell(row = 7, column = 1).font = Font(size = 12,name='Century')
        sheet.cell(row = 7, column = 1).value= 'Roll Number:'

        sheet.cell(row = 7, column = 2).alignment= Alignment('left')
        sheet.cell(row = 7, column = 2).font = Font(size = 12,bold=True,name='Century')
        sheet.cell(row = 7, column = 2).value=st_roll_list[st_no_master]

        sheet.cell(row = 10, column = 5).alignment= Alignment('center')
        sheet.cell(row = 10, column = 5).font = Font(size = 12,name='Century')
        sheet.cell(row = 10, column = 5).value= str(len(crt_options))

        i=2
        while(i<6):
            sheet.cell(row = 9, column = i).alignment= Alignment('center')
            sheet.cell(row = 9, column = i).font = Font(size = 12,bold=True,name='Century')
            if(i==2):
                x='Right'
            elif(i==3):
                x='Wrong'
            elif(i==4):
                x='Not Attempt'
            else:
                x='Max'
            sheet.cell(row = 9, column = i).value= x
            i+=1

        #marking(positive and nagative)
        sheet.cell(row = 11, column = 2).font = Font(size = 12,color='00008000',name='Century')
        sheet.cell(row = 11, column = 2).alignment= Alignment('center')
        sheet.cell(row = 11, column = 2).value=str(p)

        sheet.cell(row = 11, column = 3).font = Font(size = 12,color='00FF0000',name='Century')
        sheet.cell(row = 11, column = 3).alignment= Alignment('center')
        sheet.cell(row = 11, column = 3).value= str(n)

        sheet.cell(row = 11, column = 4).font = Font(size = 12,name='Century')
        sheet.cell(row = 11, column = 4).alignment= Alignment('center')
        sheet.cell(row = 11, column = 4).value= str(0)

        #column 15

        sheet.cell(row = 15, column = 1).alignment= Alignment('center')
        sheet.cell(row = 15, column = 1).font = Font(size = 12,bold= True,name='Century')
        sheet.cell(row = 15, column = 1).value='Student Ans'

        
        sheet.cell(row = 15, column = 4).alignment= Alignment('center')
        sheet.cell(row = 15, column = 4).font = Font(size = 12,bold= True,name='Century')
        sheet.cell(row = 15, column = 4).value='Student Ans'

        sheet.cell(row = 15, column = 2).alignment= Alignment('center')
        sheet.cell(row = 15, column = 2).font = Font(size = 12,bold= True,name='Century')
        sheet.cell(row = 15, column = 2).value='Correct Ans'

        sheet.cell(row = 15, column = 5).alignment= Alignment('center')
        sheet.cell(row = 15, column = 5).font = Font(size = 12,bold= True,name='Century')
        sheet.cell(row = 15, column = 5).value='Correct Ans'

        #inserting correct options
        line=16 
        m=0
        while(m<len(crt_options)):
            if(line<=40):

                sheet.cell(row = line, column = 2).font = Font(size = 12,color='000000FF',name='Century')
                sheet.cell(row = line, column = 2).alignment= Alignment('center')
                sheet.cell(row = line, column = 2).value= crt_options[m]

            else:
                sheet.cell(row = line-25, column =5 ).font = Font(size = 12,color='000000FF',name='Century')
                sheet.cell(row = line-25, column = 5).alignment= Alignment('center')
                sheet.cell(row = line-25, column = 5).value= crt_options[m]

            line+=1
            m+=1

        set_border(sheet,'A15:B40')
        set_border(sheet,'D15:E'+ '{}'.format(line-26))
            
        #A10,A11,A12
        i=10
        while(i<13):
            sheet.cell(row = i, column = 1).alignment= Alignment('center')
            sheet.cell(row = i, column = 1).font = Font(size = 12,bold=True,name='Century')
            if(i==10):
                x='No.'
            elif(i==11):
                x='Marking'
            else:
                x='Total'
            sheet.cell(row = i, column = 1).value= x
            i+=1

        if( st_no<len(roll_list) and st_roll_list[st_no_master]==roll_list[st_no]):

            st_options = []
            st_options = Responses_df.iloc[st_no].values.tolist()
            st_options=st_options[6:]

            

            k=0
            total=0
            wrong=0
            right=0
            notattempt=0
            while(k<len(crt_options)):
                if(st_options[k]==crt_options[k]):
                    right+=1
                k+=1

            notattempt=st_options.count(nan)
            wrong=len(st_options)-right-notattempt

            total = ((right*p)+(wrong*n))
            max_marks = len(crt_options)*p 

            

            set_border(sheet, 'A9:E12')

            # No of right or Wrong 
            sheet.cell(row = 10, column = 2).font = Font(size = 12,color='00008000',name='Century')
            sheet.cell(row = 10, column = 2).alignment= Alignment('center')
            sheet.cell(row = 10, column = 2).value=str(right)

            sheet.cell(row = 10, column = 3).font = Font(size = 12,color='00FF0000',name='Century')
            sheet.cell(row = 10, column = 3).alignment= Alignment('center')
            sheet.cell(row = 10, column = 3).value= str(wrong)

            sheet.cell(row = 10, column = 4).font = Font(size = 12,name='Century')
            sheet.cell(row = 10, column = 4).alignment= Alignment('center')
            sheet.cell(row = 10, column = 4).value= str(notattempt)



            #total

            sheet.cell(row = 12, column = 2).font = Font(size = 12,color='00008000',name='Century')
            sheet.cell(row = 12, column = 2).alignment= Alignment('center')
            sheet.cell(row = 12, column = 2).value= right*p

            sheet.cell(row = 12, column = 3).font = Font(size = 12,color='00FF0000',name='Century')
            sheet.cell(row = 12, column = 3).alignment= Alignment('center')
            sheet.cell(row = 12, column = 3).value= wrong*n

            sheet.cell(row = 12, column = 5).font = Font(size = 12,color='000000FF', name='Century')
            sheet.cell(row = 12, column = 5).alignment= Alignment('center')
            sheet.cell(row = 12, column = 5).value= '{}/{}'.format(total,max_marks)




            line=16 
            m=0
            while(m<len(crt_options)):
                if(line<=40):
                    if(st_options[m]==crt_options[m]):
                        sheet.cell(row = line, column = 1).font = Font(size = 12,color='00008000',name='Century')
                        sheet.cell(row = line, column = 1).alignment= Alignment('center')
                        sheet.cell(row = line, column = 1).value= st_options[m]
                    else:
                        sheet.cell(row = line, column = 1).font = Font(size = 12,color='00FF0000',name='Century')
                        sheet.cell(row = line, column = 1).alignment= Alignment('center')
                        sheet.cell(row = line, column = 1).value=st_options[m]

                


                else:
                    if(st_options[m]==crt_options[m]):
                            sheet.cell(row = line-25, column = 4).font = Font(size = 12,color='00008000',name='Century')
                            sheet.cell(row = line-25, column = 4).alignment= Alignment('center')
                            sheet.cell(row = line-25, column = 4).value= st_options[m]
                    else:
                            sheet.cell(row = line-25, column = 4).font = Font(size = 12,color='00FF0000',name='Century')
                            sheet.cell(row = line-25, column = 4).alignment= Alignment('center')
                            sheet.cell(row = line-25, column = 4).value=st_options[m]

                line+=1
                m+=1
            st_no+=1
        
        
        wb.save(f'my_output\\{st_roll_list[st_no_master]}.xlsx')
        st_no_master+=1

def consicesheet(p,n):
    
    current_path = os.getcwd()
    response_path= current_path+'/sample_input'
    os.chdir(response_path)
    
    Responses_df = pd.read_csv('responses.csv',index_col='Timestamp')
    roll_list = Responses_df["Roll Number"].values.tolist()
    

    st_roll_list_df=pd.read_csv('master_roll.csv')
    st_roll_list = st_roll_list_df["roll"].values.tolist()
    
    crt_options=[]
    crt_options = Responses_df.loc[Responses_df['Roll Number']=='ANSWER'].values.tolist()
    if(len(crt_options)==0):
        return True
    if(len(crt_options)==0):
        return False
    crt_options=crt_options[0]
    crt_options=crt_options[6:]

    os.chdir(current_path)
    if not os.path.exists("my_output"):
        os.mkdir('my_output')

    status=[]
    score_after_negative=[]

    st_no = 0
    st_no_master=0
    while(st_no_master < len(st_roll_list) ):    

        if( st_no<len(roll_list) and st_roll_list[st_no_master]==roll_list[st_no]):

            st_options = []
            st_options = Responses_df.iloc[st_no].values.tolist()
            st_options=st_options[6:]

            k=0
            total=0
            wrong=0
            right=0
            notattempt=0
            while(k<len(crt_options)):
                if(st_options[k]==crt_options[k]):
                    right+=1
                k+=1

            notattempt=st_options.count(nan)
            wrong=len(st_options)-right-notattempt
            status.append('[{},{},{}]'.format(right,wrong,notattempt))

            total = ((right*p)+(wrong*n))
            max_marks = len(crt_options)*p 
            score_after_negative.append('{}/{}'.format(total,max_marks))
            st_no+=1
            
        st_no_master+=1

    Responses_df.insert(5,'score_after_negative',score_after_negative,True)
    Responses_df['Status'] = status
    Responses_df.to_csv("my_output\concise_marksheet.csv")

def sendmail():

    current_path = os.getcwd()
    response_path= current_path+'/sample_input'
    os.chdir(response_path)
    
    
    Responses_df = pd.read_csv('responses.csv',index_col='Timestamp')
    
    email_list = Responses_df["Email address"].values.tolist()
    IITP_webmail_list =Responses_df["IITK webmail"].values.tolist() 
    roll_list = Responses_df["Roll Number"].values.tolist()
    dict={}
    for i in range(0,len(roll_list)):
        dict[roll_list[i]]=[IITP_webmail_list[i],email_list[i]]
    os.chdir(current_path) 
    myoutput_path = current_path+'/my_output'
    os.chdir(myoutput_path)
    for i in range(0,len(roll_list)): 
        for j in range(0,2):  
            fromaddr = "shivashruthi00@gmail.com"
            toaddr = dict[roll_list[i]][j]
            msg = MIMEMultipart() 
            msg['From'] = fromaddr 
            msg['To'] = toaddr 
            msg['Subject'] = "Quiz marks"
            body = "Your Quiz marks"
            msg.attach(MIMEText(body, 'plain')) 
            filename = "{}.xlsx".format(roll_list[i])
            attachment = open("{}.xlsx".format(roll_list[i]), "rb")
            p = MIMEBase('application', 'octet-stream')
            p.set_payload((attachment).read())
            encoders.encode_base64(p)
            p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
            msg.attach(p)
            s = smtplib.SMTP('smtp.gmail.com', 587)
            s.starttls()
            s.login(fromaddr, "Aparnashruthi")
            text = msg.as_string()
            s.sendmail(fromaddr, toaddr, text)
            s.quit()

    
    



